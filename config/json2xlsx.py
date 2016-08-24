#!/usr/bin/python3

import warnings
warnings.simplefilter("ignore")

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import *
from openpyxl.cell import get_column_letter

from io import BytesIO

import json
import sys

def main():
	data = json.load(sys.stdin)

	wb = Workbook(guess_types=False)

	st = Style(fill=PatternFill(patternType='solid', fgColor=Color('FF3b97d3')),
	           font=Font(bold=True, color=colors.WHITE))

	for sheet in data:
		if wb.active.title == "Sheet":
			ws = wb.active
		else:
			ws = wb.create_sheet()
		ws.title = sheet['name']
		widths = {}
		ws.append(sheet['header'])
		for idx, val in enumerate(sheet['header']):
			widths[idx] = len(val)

		if not sheet['rows'] is None:
			for row in sheet['rows']:
				ws.append(row)
				for idx, val in enumerate(row):
					try:
						ln = len(str(val))
					except:
						ln = 10
					if idx not in widths or ln > widths[idx]:
						widths[idx] = ln

		for k in widths:
			ws.column_dimensions[get_column_letter(k+1)].width = widths[k] + 2
		
		for col_idx in range(1, 26):
			col = get_column_letter(col_idx)
			ws.cell('%s%s'%(col, 1)).style = st

	out = BytesIO()
	wb.save(filename = out)
	out.seek(0)
	if hasattr(sys.stdout, 'buffer'):
		sys.stdout.buffer.write(out.read())
	else: 
		sys.stdout.write(out.read())
	sys.stdout.close()

if __name__ == '__main__':
	main()
